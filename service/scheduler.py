import argparse
import copy
import datetime
import os

import openpyxl

from model.AssignmentCount import AssignmentCount


class ValidateInt(argparse.Action):
    def __call__(self, parser, namespace, values, option_string=None):
        for value in values:
            try:
                int(value)
            except ValueError:
                parser.error("请输入一个数字，获取到的参数为{}".format(value))
        setattr(namespace, self.dest, values)

dir_prefix = ".."

tool_description = """
排班程序说明：
"""
parser = argparse.ArgumentParser(description=tool_description)
parser.add_argument("-c", "--calendar", dest="calendar", help="生成日历, 指定年份, 如2022")
parser.add_argument("-m", "--month", dest="month", help="排班月份，格式为yyyymm, 如202209")
parser.add_argument("-s", "--stats", dest="stats", help="统计历史值班数据, 无需参数", action="store_true")


class Assignment(object):

    def __init__(self, date: datetime.date | str, name: str, holiday: bool = False):
        self.date = date if isinstance(date, datetime.date) else datetime.datetime.strptime(date, "%Y-%m-%d").date()
        self.name = name
        self.holiday = holiday


class Stat(object):

    def __init__(self, name: str, stats: float):
        self.name = name
        self.stats = stats


class WorkDay(object):

    def __init__(self, date: str, holiday: bool):
        self.date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
        self.holiday = holiday

    def __str__(self):
        return "{} {}".format(self.date, self.holiday)

    def __repr__(self):
        return self.__str__()


class Scheduler(object):

    schedule_dir = os.path.join(dir_prefix, "../schedule")
    workday_dir = os.path.join(dir_prefix, "../schedule", "workdays")
    worker_dir = os.path.join(dir_prefix, "../schedule", "workers.xlsx")
    history_dir = os.path.join(dir_prefix, "../schedule", "history")
    dest_dir = os.path.join(dir_prefix, "../schedule", "this_year")
    Y = "Y"
    N = "N"
    NORMAL_WORKDAY = 1
    LAST_WORKDAY = 2
    NORMAL_HOLIDAY = 3
    LAST_HOLIDAY = 4
    weekday_name = {
        1: "周一",
        2: "周二",
        3: "周三",
        4: "周四",
        5: "周五",
        6: "周六",
        7: "周天"
    }
    assignment_header = ["日期", "星期几", "值班人员", "是否节假日"]
    stats_header = ["姓名", "普通工作日", "节前工作日", "工作日值班总数", "节前工作日排班系数", "普通节假日", "节假日最后一天", "节假日值班总数", "普通节假日排班系数"]

    def __init__(self, __print):
        self.print = __print
        self.normal_workdays: dict[int: dict[int: list[int]]] = {}
        self.last_workdays: dict[int: dict[int: list[int]]] = {}
        self.normal_holidays: dict[int: dict[int: list[int]]] = {}
        self.last_holidays: dict[int: dict[int: list[int]]] = {}
        self.workers: list[str] = []
        # last workday stats
        self.workday_stats: list[Stat] = []
        self.workday_stats_map: dict[str: float] = {}
        # normal holiday stats
        self.holiday_stats: list[Stat] = []
        self.holiday_stats_map: dict[str: float] = {}
        self.last_workday_dest: float = 20.0
        self.normal_holiday_dest: float = 50.0
        self.history_schedule: list[Assignment] = []
        self.history_count: dict[str: AssignmentCount] = {}
        self.final_count: dict[str: AssignmentCount] = {}

    def check_and_create_dir(self):
        if not os.path.exists(self.schedule_dir):
            self.print("{} 不存在, 创建中...".format(self.schedule_dir))
            os.mkdir(self.schedule_dir)
        if not os.path.exists(self.workday_dir):
            self.print("{} 不存在, 创建中...".format(self.workday_dir))
            os.mkdir(self.workday_dir)
        if not os.path.exists(self.history_dir):
            self.print("{} 不存在, 创建中...".format(self.history_dir))
            os.mkdir(self.history_dir)
        if not os.path.exists(self.dest_dir):
            self.print("{} 不存在, 创建中...".format(self.dest_dir))
            os.mkdir(self.dest_dir)

    def load_history(self):
        self.print("开始加载历史数据")
        history_files = self._get_files(self.history_dir)
        for file_name in history_files:
            self.print("加载历史数据文件: {}".format(file_name))
            wb = openpyxl.load_workbook(filename=file_name, data_only=True)
            for sheet in wb.worksheets:
                rows = list(sheet.rows)[1:]
                for row in rows:
                    self.history_schedule.append(
                        Assignment(row[0].value, row[2].value, str(row[3].value).upper() == self.Y))
        self.history_schedule.sort(key=lambda x: x.date)
        self.print("加载完成!")

    def load_days(self):
        self.print("开始加载日历文件")
        workday_files = self._get_files(self.workday_dir)
        days = []
        for file_name in workday_files:
            wb = openpyxl.load_workbook(filename=file_name, data_only=True)
            for sheet in wb.worksheets:
                rows = list(sheet.rows)[1:]
                for row in rows:
                    days.append(WorkDay(row[0].value, self.Y == str(row[2].value).upper()))

        days.sort(key=lambda x: x.date)

        for i in range(len(days)):
            next_holiday = None
            if i + 1 < len(days):
                next_holiday = days[i+1].holiday
            if next_holiday is None:
                # 没有next day的情况应该是12月31日，1月1日肯定为假期，所以可以这样处理
                if days[i].holiday:
                    self._put_to_day(self.normal_holidays, days[i])
                else:
                    self._put_to_day(self.last_workdays, days[i])
                continue

            if days[i].holiday and next_holiday:
                # normal holiday
                self._put_to_day(self.normal_holidays, days[i])
            elif days[i].holiday and not next_holiday:
                # last holiday
                self._put_to_day(self.last_holidays, days[i])
            elif not days[i].holiday and next_holiday:
                # last workday
                self._put_to_day(self.last_workdays, days[i])
            else:
                # normal workday
                self._put_to_day(self.normal_workdays, days[i])
        self.print("加载完成!")

    def _put_to_day(self, days: dict[int: dict[int, list[int]]], work_day: WorkDay):
        year = work_day.date.year
        month = work_day.date.month
        day = work_day.date.day
        if year not in days:
            days[year] = {i+1: [] for i in range(12)}
        days[year][month].append(day)

    def _get_files(self, dir: str) -> list[str]:
        files = []
        for f in os.listdir(dir):
            if os.path.isfile(os.path.join(dir, f)):
                files.append(os.path.join(dir, f))
        return files

    def generate_days(self, year=None):
        self.print("正在创建日历...")
        if year is None:
            next_year = datetime.datetime.now() + datetime.timedelta(days=365)
            year = next_year.year
        days = []
        cur = datetime.date(year=year, month=1, day=1)
        while cur.year == year:
            days.append(cur)
            cur = cur + datetime.timedelta(days=1)

        sheet_names = ["{}月".format(i + 1) for i in range(12)]
        data = {}
        for name in sheet_names:
            data[name] = [["日期", "星期名称", "节假日"]]

        for day in days:
            d = day.strftime("%Y-%m-%d")
            data["{}月".format(day.month)].append([d, day.isoweekday(), self.Y if day.isoweekday() > 5 else self.N])

        path = os.path.join(self.workday_dir, "{}.xlsx".format(year))
        self._save_to_excel(path, sheet_names, data)
        self.print("创建完成")
        return path

    def _save_to_excel(self, path: str, sheet_names: list[str], data: dict[str: list[list[any]]],
                       workbook: openpyxl.Workbook | None = None):
        wb: openpyxl.Workbook | None = None
        if workbook:
            wb = workbook
            for name in sheet_names:
                if name in wb:
                    self.print("表单{}已经存在，默认清空".format(name))
                    wb.remove(wb[name])
                wb.create_sheet(title=name)
        else:
            wb = openpyxl.Workbook()
            wb.active.title = sheet_names[0]
            for name in sheet_names[1:]:
                wb.create_sheet(title=name)

        for sheet in wb.worksheets:
            sheet_data = data.get(sheet.title)
            if sheet_data:
                for row in sheet_data:
                    sheet.append(row)
        wb.save(filename=path)

    def load_workers(self):
        wb = openpyxl.load_workbook(self.worker_dir, data_only=True)
        ws = wb.active
        rows = list(ws.rows)[1:]
        self.workers = []
        for row in rows:
            name = row[0].value
            need_schedule = row[1].value
            if str(need_schedule) == self.Y:
                self.workers.append(name)

    def stats_history(self) -> str:
        self.load_days()
        self.load_history()
        self._final_stats([])
        file_name = self._save_assignment_stat(None, None)
        self.print("统计完成, 统计文件位于: ", file_name)
        return file_name

    def schedule_for_month(self, year: int, month: int) -> (str, str):
        self.load_workers()
        self.load_days()
        self.load_history()
        self._init_stats()
        self.print("开始进行排班规划")
        workday_schedules = self.schedule_workday(year, month)
        holiday_schedules = self.schedule_holiday(year, month)
        schedules = workday_schedules + holiday_schedules
        schedules.sort(key=lambda x: x.date)
        self._reschedule_continuous(schedules)
        self._final_stats(schedules)
        file_name = self._save_schedule(year, month, schedules)
        self.print("排班完成, 排班文件位于: ", file_name)
        stats_file_name = self._save_assignment_stat(year, month)
        self.print("统计完成, 统计文件位于: ", stats_file_name)
        return (file_name, stats_file_name)

    def _save_schedule(self, year: int, month: int, schedules: list[Assignment]) -> str:
        file_name = os.path.join(self.dest_dir, "{}年{}月值班安排.xlsx".format(year, month))
        wb = openpyxl.load_workbook(file_name, data_only=True) if os.path.exists(file_name) else None
        name = "{}月".format(month)
        data = {name: [self.assignment_header]}
        rows = data[name]
        for s in schedules:
            holiday_flag = self.Y if s.holiday else self.N
            rows.append([s.date.strftime("%Y-%m-%d"), self.weekday_name[s.date.isoweekday()], s.name, holiday_flag])

        self._save_to_excel(file_name, [name], data, wb)
        return file_name

    def _save_assignment_stat(self, year: int, month: int = None) -> str:
        if month is not None:
            file_name = os.path.join(self.dest_dir, "{}年{}月值班统计.xlsx".format(year, month))
        else:
            file_name = os.path.join(self.dest_dir, "历史统计.xlsx")
        wb = openpyxl.load_workbook(file_name, data_only=True) if os.path.exists(file_name) else None
        sheetname = "历史" if month is None else "{}月".format(month)
        data = {sheetname: [self.stats_header]}
        rows = data[sheetname]
        for name, count in self.final_count.items():
            rows.append([name, count.normal_workday, count.last_workday, count.normal_workday + count.last_workday,
                         self.workday_stats_map.get(name, 0.0), count.normal_holiday, count.last_holiday,
                         count.normal_holiday +count.last_holiday, self.holiday_stats_map.get(name, 0.0)])
        self._save_to_excel(file_name, [sheetname], data, wb)
        return file_name

    def schedule_workday(self, year: int, month: int) -> list[Assignment]:
        # 避免名单中在前面的倒霉蛋排太多值班
        total_count = {x: y.workday for x, y in self.history_count.items()}
        self._resort_person(total_count)
        start_idx = self._find_start_person_idx(lambda x: not x.holiday)
        schedules = self._schedule(year, month, start_idx, self.last_workdays[year][month],
                                   self.normal_workdays[year][month], self.workday_stats, self.workday_stats_map,
                                   self.last_workday_dest)
        return schedules

    def schedule_holiday(self, year: int, month: int) -> list[Assignment]:
        total_count = {x: y.holiday for x, y in self.history_count.items()}
        self._resort_person(total_count)
        start_idx = self._find_start_person_idx(lambda x: x.holiday)
        schedules = self._schedule(year, month, start_idx, self.normal_holidays[year][month],
                                   self.last_holidays[year][month], self.holiday_stats, self.holiday_stats_map,
                                   self.normal_holiday_dest)
        for s in schedules:
            s.holiday = True
        return schedules

    def _resort_person(self, total_count: dict[str: int]):
        self.workers.sort(key=lambda y: total_count.get(y, 0))

    def _find_start_person_idx(self, predict) -> int:
        for i in range(len(self.history_schedule)-1, 0, -1):
            if predict(self.history_schedule[i]):
                for j in range(len(self.workers)):
                    if self.workers[j] == self.history_schedule[i].name:
                        return j + 1
        return 0

    def _schedule(self, year: int, month: int, idx: int, stats_day: list[int], non_stats_day: list[int],
                  stats: list[Stat], stats_map: dict[str, float], dest: float) -> list[Assignment]:
        schedules = self._init_schedule(year, month, idx, stats_day + non_stats_day)
        self._recalc_stats(schedules)
        schedules = self._reschedule_by_stat(schedules, stats, stats_map, stats_day, dest)
        return schedules

    def _init_schedule(self, year: int, month: int, idx: int, days: list[int]) -> list[Assignment]:
        """先按顺序安排所有人"""
        cur = datetime.date(year=year, month=month, day=1)
        schedules: list[Assignment] = []
        while cur.month == month:
            worker = self.workers[idx % len(self.workers)]
            if cur.day in days:
                schedules.append(Assignment(cur, worker))
                idx += 1
            cur = cur + datetime.timedelta(days=1)
        return schedules

    def _reschedule_continuous(self, schedules: list[Assignment]):
        for i in range(len(schedules) - 1):
            if schedules[i].name == schedules[i+1].name:
                dest = schedules[i]
                dest_kind = self._get_day_kind(schedules[i].date.year, schedules[i].date.month, schedules[i].date.day)
                for j in range(i+1, len(schedules)):
                    candidate = schedules[j]
                    candi_kind = self._get_day_kind(candidate.date.year, candidate.date.month, candidate.date.day)
                    if dest.date.weekday() == candidate.date.weekday() and dest_kind == candi_kind:
                        dest.name, candidate.name = candidate.name, dest.name

    def _get_day_kind(self, year: int, month: int, day: int) -> int:
        if day in self.normal_workdays[year][month]:
            return self.NORMAL_WORKDAY
        elif day in self.last_workdays[year][month]:
            return self.LAST_WORKDAY
        elif day in self.normal_holidays[year][month]:
            return self.NORMAL_HOLIDAY
        else:
            return self.LAST_HOLIDAY

    def _reschedule_by_stat(self, schedules: list[Assignment], stats: list[Stat], stats_map: dict[str: float],
                            stats_day: list[int], dest: float) -> list[Assignment]:
        """按照统计对人员进行重新分配"""
        for i in range(len(schedules)):
            cur_worker_stats = stats_map.get(schedules[i].name, 0.0)
            day = schedules[i].date.day
            if cur_worker_stats > dest and day in stats_day:
                self._find_and_reschedule_by_stat(i, cur_worker_stats, schedules, stats, stats_day)
        return schedules

    def _find_and_reschedule_by_stat(self, dest_idx: int, dest_stats: float, schedules: list[Assignment], stats: list[Stat],
                                     stats_day: list[int]) -> None:
        """寻找可替换的人并进行调换"""
        cur_schedule = schedules[dest_idx]
        for j in range(len(stats)):
            if stats[j].stats < dest_stats:
                candidate = stats[0].name
                candidate_scheduled_flag = False
                for candi_schedule in schedules:
                    if candi_schedule.name == candidate and candi_schedule.date.day in stats_day:
                        candi_schedule.name = cur_schedule.name
                        cur_schedule.name = candidate
                        self._recalc_stats(schedules)
                        return
                else:
                    # The candidate doesn't be scheduled in this month
                    if candidate_scheduled_flag is False:
                        cur_schedule.name = candidate
                        self._recalc_stats(schedules)
                        return
            else:
                break

    def _recalc_stats(self, schedules: list[Assignment] | None):
        counts: dict[str: AssignmentCount] = copy.deepcopy(self.history_count)
        counts = self._schedule_count(schedules, counts)
        self._calc_stats(counts)

    def _final_stats(self, schedules: list[Assignment]):
        total_schedules = self.history_schedule + schedules
        self.final_count = self._schedule_count(total_schedules)
        self._calc_stats(self.final_count)

    def _init_stats(self):
        self.history_count = self._schedule_count(self.history_schedule)
        self._calc_stats(self.history_count)

    def _schedule_count(self, schedules: list[Assignment], counts: dict[str: AssignmentCount] = None) -> dict[str: AssignmentCount]:
        if counts is None:
            counts = dict()

        for s in schedules:
            year = s.date.year
            month = s.date.month
            day = s.date.day
            worker = s.name
            assignment_count = counts.get(worker)
            if assignment_count is None:
                assignment_count = AssignmentCount(worker)
                counts[worker] = assignment_count
            if day in self.normal_holidays[year][month]:
                assignment_count.normal_holiday += 1
            elif day in self.last_workdays[year][month]:
                assignment_count.last_workday += 1
            elif day in self.normal_workdays[year][month]:
                assignment_count.normal_workday += 1
            else:
                assignment_count.last_holiday += 1
        return counts

    def _calc_stats(self, counts: dict[str, AssignmentCount]):
        """更新stats map, stats map是混用的，这里是个天坑"""
        self.workday_stats.clear()
        self.holiday_stats.clear()
        for name, count in counts.items():
            if count.normal_workday + count.last_workday == 0:
                wstats = 0.0
            else:
                wstats = count.last_workday / (count.normal_workday + count.last_workday) * 100.0

            if count.normal_holiday + count.last_holiday == 0:
                hstats = 0.0
            else:
                hstats = count.normal_holiday / (count.normal_holiday + count.last_holiday) * 100.0
            self.workday_stats.append(Stat(name, wstats))
            self.holiday_stats.append(Stat(name, hstats))
        self.workday_stats.sort(key=lambda x: x.stats)
        self.holiday_stats.sort(key=lambda x: x.stats)

        self.workday_stats_map = {x.name: x.stats for x in self.workday_stats}
        self.holiday_stats_map = {x.name: x.stats for x in self.holiday_stats}


if __name__ == '__main__':
    # scheduler = Scheduler()
    # scheduler.schedule_for_month(2022, 1)
    # scheduler.schedule_for_month(2022, 2)
    # scheduler.schedule_for_month(2022, 3)
    # scheduler.schedule_for_month(2022, 4)
    # scheduler.schedule_for_month(2022, 5)
    # scheduler.schedule_for_month(2022, 6)
    # scheduler.schedule_for_month(2022, 7)
    # scheduler.schedule_for_month(2022, 8)
    # scheduler.schedule_for_month(2022, 9)
    # scheduler.load_days()
    # scheduler.load_workers()
    # self.print(scheduler.workers)
    args = parser.parse_args()
    scheduler = Scheduler(print)
    scheduler.check_and_create_dir()
    if args.calendar:
        scheduler.generate_days(int(args.calendar))
    elif args.stats:
        scheduler.stats_history()
    else:
        ym = int(args.month)
        year = ym // 100
        month = ym % 100
        scheduler.schedule_for_month(year, month)
