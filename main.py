import copy
import datetime
import os

import openpyxl

dir_prefix = "."


class Assignment(object):

    def __init__(self, date: datetime.date, name: str):
        self.date = date
        self.name = name


class Stat(object):

    def __init__(self, name: str, stats: float):
        self.name = name
        self.stats = stats


class WorkDay(object):

    def __init__(self, date: str, holiday: bool):
        self.date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
        self.holiday = holiday

    def __str__(self):
        return "{} {}".format(self.date, self.holiday)

    def __repr__(self):
        return self.__str__()


class AssignmentCount(object):

    def __init__(self, name):
        self.name = name
        self.normal_workday = 0
        self.last_workday = 0
        self.normal_holiday = 0
        self.last_holiday = 0


class Scheduler(object):

    workday_dir = os.path.join(dir_prefix, "workdays")
    worker_dir = os.path.join(dir_prefix, "schedule", "workers.xlsx")
    history_dir = os.path.join(dir_prefix, "schedule", "history")
    dest_dir = os.path.join(dir_prefix, "schedule", "this_year")
    Y = "Y"
    N = "N"

    def __init__(self):
        self.normal_workdays: dict[int: dict[int: list[int]]] = {}
        self.last_workdays: dict[int: dict[int: list[int]]] = {}
        self.normal_holidays: dict[int: dict[int: list[int]]] = {}
        self.last_holidays: dict[int: dict[int: list[int]]] = {}
        self.workers: list[str] = []
        # last workday stats
        self.workday_stats: list[Stat] = []
        self.workday_stats_map: dict[str: float] = {}
        # normal holiday stats
        self.holiday_stats: list[Stat] = []
        self.holiday_stats_map: dict[str: float] = {}
        self.last_workday_dest: float = 20.0
        self.normal_holiday_dest: float = 50.0
        self.history_schedule: list[Assignment] = []
        self.history_count: dict[str: AssignmentCount] = {}

    def load_history(self):
        pass

    def load_days(self):
        workday_files = self._get_files(self.workday_dir)
        days = []
        for file_name in workday_files:
            wb = openpyxl.load_workbook(filename=file_name, data_only=True)
            for sheet in wb.worksheets:
                rows = list(sheet.rows)[1:]
                for row in rows:
                    days.append(WorkDay(row[0].value, self.Y == str(row[2].value).upper()))

        for i in range(len(days)):
            next_holiday = None
            if i + 1 < len(days):
                next_holiday = days[i+1].holiday
            if next_holiday is None:
                # 没有next day的情况应该是12月31日，1月1日肯定为假期，所以可以这样处理
                if days[i].holiday:
                    self._put_to_day(self.normal_holidays, days[i])
                else:
                    self._put_to_day(self.last_workdays, days[i])
                continue

            if days[i].holiday and next_holiday:
                # normal holiday
                self._put_to_day(self.normal_holidays, days[i])
            elif days[i].holiday and not next_holiday:
                # last holiday
                self._put_to_day(self.last_holidays, days[i])
            elif not days[i].holiday and next_holiday:
                # last workday
                self._put_to_day(self.last_workdays, days[i])
            else:
                # normal workday
                self._put_to_day(self.normal_workdays, days[i])

    def _put_to_day(self, days: dict[int: dict[int, list[int]]], work_day: WorkDay):
        year = work_day.date.year
        month = work_day.date.month
        day = work_day.date.day
        if year not in days:
            days[year] = {i+1: [] for i in range(12)}
        days[year][month].append(day)

    def _get_files(self, dir: str) -> list[str]:
        files = []
        for f in os.listdir(dir):
            if os.path.isfile(os.path.join(dir, f)):
                files.append(os.path.join(dir, f))
        return files

    def generate_days(self, year=None):
        if year is None:
            next_year = datetime.datetime.now() + datetime.timedelta(days=365)
            year = next_year.year
        days = []
        cur = datetime.date(year=year, month=1, day=1)
        while cur.year == year:
            days.append(cur)
            cur = cur + datetime.timedelta(days=1)

        sheet_names = ["{}月".format(i + 1) for i in range(12)]
        data = {}
        for name in sheet_names:
            data[name] = [["日期", "星期名称", "节假日"]]

        for day in days:
            d = "{}-{}-{}".format(day.year, day.month, day.day)
            data["{}月".format(day.month)].append([d, day.isoweekday(), self.Y if day.isoweekday() > 5 else self.N])

        path = os.path.join(self.workday_dir, "{}.xlsx".format(year))
        self._save_to_excel(path, sheet_names, data)

    def _save_to_excel(self, path: str, sheet_names: list[str], data: dict[str: list[list[any]]]):

        wb = openpyxl.Workbook()
        wb.active.title = sheet_names[0]
        for name in sheet_names[1:]:
            wb.create_sheet(title=name)

        for sheet in wb.worksheets:
            sheet_data = data.get(sheet.title)
            if sheet_data:
                for row in sheet_data:
                    sheet.append(row)
        wb.save(filename=path)

    def load_workers(self):
        wb = openpyxl.load_workbook(self.worker_dir, data_only=True)
        ws = wb.active
        rows = list(ws.rows)[1:]
        self.workers = []
        for row in rows:
            name = row[0].value
            need_schedule = row[1].value
            if str(need_schedule) == self.Y:
                self.workers.append(name)

    def schedule_for_month(self, year: int, month: int) -> list[Assignment]:
        self.load_workers()
        self.load_days()
        self.load_history()
        self._calc_stats()
        workday_schedules = self.schedule_workday(year, month)
        holiday_schedules = self.schedule_holiday(year, month)
        schedules = workday_schedules + holiday_schedules
        return sorted(schedules, key=lambda x: x.date)

    def schedule_workday(self, year: int, month: int) -> list[Assignment]:
        schedules = self._schedule(year, month, self.last_workdays[year][month], self.normal_workdays[year][month],
                                   self.workday_stats, self.workday_stats_map, self.last_workday_dest)
        # TODO: nobody will be scheduled two times in a week is preferred
        return schedules

    def schedule_holiday(self, year: int, month: int) -> list[Assignment]:
        schedules = self._schedule(year, month, self.normal_holidays[year][month], self.last_holidays[year][month],
                                   self.holiday_stats, self.holiday_stats_map, self.normal_holiday_dest)
        return schedules

    def _schedule(self, year: int, month: int, stats_day: list[int], non_stats_day: list[int], stats: list[Stat],
                  stats_map: dict[str, float], dest: float) -> list[Assignment]:
        schedules = self._init_schedule(year, month, stats_day + non_stats_day)
        schedules = self._reschedule_by_stat(schedules, stats, stats_map, stats_day, dest)
        return schedules

    def _init_schedule(self, year: int, month: int, days: list[int]) -> list[Assignment]:
        """先按顺序安排所有人"""
        cur = datetime.date(year=year, month=month, day=1)
        count = 0
        schedules: list[Assignment] = []
        while cur.month == month:
            worker = self.workers[count % len(self.workers)]
            if cur.day in days:
                schedules.append(Assignment(cur, worker))
                count += 1
            cur = cur + datetime.timedelta(days=1)
        return schedules

    def _reschedule_by_stat(self, schedules: list[Assignment], stats: list[Stat], stats_map: dict[str: float],
                            stats_day: list[int], dest: float) -> list[Assignment]:
        """按照统计对人员进行重新分配"""
        for i in range(len(schedules)):
            cur_worker_stats = stats_map.get(schedules[i].name, 0.0)
            if cur_worker_stats > dest:
                self._reschedule(i, cur_worker_stats, schedules, stats, stats_day)
        return schedules

    def _reschedule(self, dest_idx: int, dest_stats: float, schedules: list[Assignment], stats: list[Stat],
                    stats_day: list[int]) -> None:
        """寻找可替换的人并进行调换"""
        cur_schedule = schedules[dest_idx]
        for j in range(len(stats)):
            if stats[j].stats < dest_stats:
                candidate = stats[0].name
                candidate_scheduled_flag = False
                for candi_schedule in schedules:
                    if candi_schedule.name == candidate and candi_schedule.date.day in stats_day:
                        candi_schedule.name = cur_schedule.name
                        cur_schedule.name = candidate
                        self._recalc_stats(schedules)
                        return
                else:
                    # The candidate doesn't be scheduled in this month
                    if candidate_scheduled_flag is False:
                        cur_schedule.name = candidate
                        self._recalc_stats(schedules)
                        return
            else:
                break

    def _recalc_stats(self, schedules: list[Assignment] | None):
        counts: dict[str: AssignmentCount] = copy.deepcopy(self.history_count)
        for s in schedules:
            year = s.date.year
            month = s.date.month
            day = s.date.day
            worker = s.name
            assignment_count = counts.get(worker)
            if assignment_count is None:
                assignment_count = AssignmentCount(worker)
                counts[worker] = assignment_count
            if day in self.normal_holidays[year][month]:
                assignment_count.normal_workday += 1
            elif day in self.last_workdays[year][month]:
                assignment_count.last_workday += 1
            elif day in self.normal_workdays[year][month]:
                assignment_count.normal_holiday += 1
            else:
                assignment_count.last_holiday += 1

        for ws in self.workday_stats:
            count: AssignmentCount = counts[ws.name]
            ws.stats = count.last_workday / (count.normal_workday + count.last_workday) * 100
            self.workday_stats_map[ws.name] = ws.stats
        self.workday_stats.sort(key=lambda x: x.date)

        for hs in self.holiday_stats:
            count: AssignmentCount = counts[hs.name]
            hs.stats = count.normal_holiday / (count.normal_holiday + count.last_holiday) * 100
            self.holiday_stats_map[hs.name] = hs.stats
        self.holiday_stats.sort(key=lambda x: x.date)

    def _calc_stats(self):
        self._recalc_stats(self.history_schedule)


if __name__ == '__main__':
    scheduler = Scheduler()
    scheduler.schedule_for_month(2022, 7)
    # scheduler.load_days()
    # scheduler.load_workers()
    # print(scheduler.workers)
